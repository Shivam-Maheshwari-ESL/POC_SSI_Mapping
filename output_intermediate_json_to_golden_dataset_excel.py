import json
import os
import re
import glob
from copy import copy
from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================


INPUT_EXCEL = r"SSI_Golden_Set_Expected_Extraction_Output_mapped.xlsx"

JSON_FOLDER = r"output_llm_on_ADI\intermediate_llm_json"

OUTPUT_EXCEL = r"SSI_Golden_Set_Expected_Extraction_Output_mapped.xlsx"

# Filename pattern used to extract the document key from each JSON file,
# e.g. "output_output_ADI_2_gpt-5.4_intermediate.json" -> "2"
JSON_NAME_PATTERN = re.compile(
    r"^output_output_ADI_(.+)_gpt-5\.4_intermediate\.json$"
)


# ============================================================
# STANDARD EXCEL COLUMN POSITIONS
# ============================================================

COMMON_COLUMNS = {
    "action": 2,
    "instructingParty.identifier": 3,
    "instructingParty.lei": 4,
    "instructingParty.bic": 5,
    "instructingParty.proprietaryIdentification": 6,
    "instructingParty.proprietarySchemeName": 7,
    "instructingParty.proprietaryIssuer": 8,

    "beneficiary.entityIdentification": 10,

    "depository.account": 11,
    "depository.accountIdentification": 12,

    "market.identifier2": 13,
    "market.identification2": 14,
    "market.identifier3": 15,
    "market.identification3": 16,
    "market.identifier4": 17,
    "market.identification4": 18,
    "market.identifier5": 19,
    "market.identification5": 20,

    "effectiveDate.parameter": 21,
    "effectiveDate.start": 22,
    "effectiveDate.end": 23,

    "market.settlementPurpose": 24,
    "market.narrative": 25,
    "market.settlementCountry": 26,
    "market.classificationIdentifier": 27,
    "market.cfi": 28,
    "market.alternateClassificationIdentification": 29,
    "market.alternateSchemeName": 30,
    "market.alternateIssuer": 31,
    "market.isin": 32,
    "market.settlementCurrency": 33,
    "market.psafeBic": 34,
    "market.isinCountry": 35,

    "depository.psetBic": 36,
}


# Party 1 starts at column 37.
# Each party occupies 7 columns.
PARTY_START_COLUMN = 37
PARTY_COLUMN_WIDTH = 7

PARTY_FIELDS = {
    "identifier": 0,
    "bic": 1,
    "proprietaryIdentification": 2,
    "proprietarySchemeName": 3,
    "proprietaryIssuer": 4,
    "accountIdentification": 5,
    "accountName": 6,
}


# Some intermediate JSON files use plain numeric fieldIds (e.g. "15")
# instead of the dotted key (e.g. "market.settlementCurrency"). This table
# was derived by cross-referencing fieldName <-> fieldId pairs across all
# JSON files in output_llm_on_ADI/intermediate_llm_json, where every numeric
# fieldId has an exact, conflict-free dotted-key counterpart under the same
# fieldName in other files. It is used only to resolve which column a field
# belongs to, never to invent or alter any mappedValue.
LABEL_TO_FIELD_ID = {
    "Action": "action",
    "Instructing Party Identifier": "instructingParty.identifier",
    "Instructing Entity LEI": "instructingParty.lei",
    "Instructing Party BIC": "instructingParty.bic",
    "Proprietary Identification": "instructingParty.proprietaryIdentification",
    "Proprietary Scheme Name": "instructingParty.proprietarySchemeName",
    "Beneficiary Entity Identification": "beneficiary.entityIdentification",
    "Effective Date Parameter": "effectiveDate.parameter",
    "Effective Start Date": "effectiveDate.start",
    "Effective End Date": "effectiveDate.end",
    "Settlement Purpose": "market.settlementPurpose",
    "Settlement Country": "market.settlementCountry",
    "Classification Identifier": "market.classificationIdentifier",
    "ISIN": "market.isin",
    "Settlement Currency": "market.settlementCurrency",
    "PSAFE BIC": "market.psafeBic",
    "PSET Party Identifier - BIC": "depository.psetBic",
    "Party 1 Party Identifier": "party1.identifier",
    "Party 1 BIC": "party1.bic",
    "Party 1 Account Identification": "party1.accountIdentification",
    "Party 1 Account Name": "party1.accountName",
    "Party 2 Party Identifier": "party2.identifier",
    "Party 2 BIC": "party2.bic",
    "Party 2 Account Identification": "party2.accountIdentification",
    "Party 2 Account Name": "party2.accountName",
    "Party 3 Party Identifier": "party3.identifier",
    "Party 3 BIC": "party3.bic",
    "Party 3 Account Identification": "party3.accountIdentification",
    "Party 3 Account Name": "party3.accountName",
    "Party 4 Party Identifier": "party4.identifier",
    "Party 4 BIC": "party4.bic",
    "Party 4 Account Identification": "party4.accountIdentification",
    "Party 4 Account Name": "party4.accountName",
}


# ============================================================
# HELPERS
# ============================================================

def party_column(party_number, field_name):
    """
    Return Excel column for a party field.

    Party 1:
        37-43

    Party 2:
        44-50

    Party 3:
        51-57

    ...
    """
    start = PARTY_START_COLUMN + (party_number - 1) * PARTY_COLUMN_WIDTH
    return start + PARTY_FIELDS[field_name]


def excel_column_for_field(field_name):
    """
    Convert JSON field name to Excel column number.
    """

    # Common fields
    if field_name in COMMON_COLUMNS:
        return COMMON_COLUMNS[field_name]

    # Party fields
    if field_name.startswith("party"):
        parts = field_name.split(".", 1)

        if len(parts) != 2:
            return None

        party_part, sub_field = parts

        try:
            party_number = int(party_part.replace("party", ""))
        except ValueError:
            return None

        if sub_field not in PARTY_FIELDS:
            return None

        return party_column(party_number, sub_field)

    return None


def clear_data_rows(ws, start_row=4):
    """
    Clear values from row 4 onward while preserving
    formatting, borders, widths, etc.
    """

    for row in ws.iter_rows(
        min_row=start_row,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column
    ):
        for cell in row:
            cell.value = None


def copy_row_format(ws, source_row, target_row):
    """
    Copy formatting from source row to target row.

    This is useful if the target workbook contains
    formatting only on the first data row.
    """

    for col in range(1, ws.max_column + 1):

        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)

        if source.has_style:
            target._style = copy(source._style)

        if source.number_format:
            target.number_format = source.number_format

        if source.alignment:
            target.alignment = copy(source.alignment)

        if source.protection:
            target.protection = copy(source.protection)


def get_records_from_json(data):
    """
    Get the primary records from the JSON.
    """

    records = data.get("records", [])

    if not isinstance(records, list):
        return []

    return records


def build_review_record_map(data):
    """
    Build a map of recordId -> fields from unmappedAndReview.

    Some of the SSI JSON outputs have:

        records[].fields[].mappedValue = null

    while the actual standardized values appear under:

        unmappedAndReview[].fields

    when a recordId is supplied.

    This helper supports that structure without
    guessing values from sourceValues/evidence.
    """

    review_map = {}

    for item in data.get("unmappedAndReview", []):

        record_id = item.get("recordId")

        if not record_id:
            continue

        fields = item.get("fields", [])

        if isinstance(fields, list):
            review_map[record_id] = fields

    return review_map


def get_field_value(field):
    """
    Return ONLY mappedValue.

    Never fall back to sourceValues or evidence.
    """

    return field.get("mappedValue")


def sheet_base_name(sheet_name):
    """
    Reduce a sheet name to its document key for matching against
    JSON file keys, e.g. "2.Pdf" -> "2", "4." -> "4".
    """

    base = sheet_name.split(".", 1)[0]

    return base.strip().lower()


def find_matching_sheet(doc_key, sheet_names):
    """
    Find the sheet whose base name matches the JSON document key.
    """

    doc_key = doc_key.strip().lower()

    for sheet_name in sheet_names:
        if sheet_base_name(sheet_name) == doc_key:
            return sheet_name

    return None


def map_records_to_sheet(ws, records):
    """
    Clear a sheet's data rows and populate it from JSON records.
    Returns the number of values mapped.
    """

    clear_data_rows(ws, start_row=4)

    total_mapped = 0

    for record_index, record in enumerate(records):

        excel_row = 4 + record_index

        # Preserve first data-row formatting
        if excel_row > 4:
            copy_row_format(ws, 4, excel_row)

        fields = record.get("fields", [])

        for field in fields:

            # fieldId is normally the machine key (e.g. "instructingParty.identifier"),
            # but some files use a plain numeric fieldId with the real key
            # only recoverable via the exact fieldName label.
            field_id = field.get("fieldId")

            if not field_id:
                continue

            resolved_field_id = field_id

            if excel_column_for_field(field_id) is None:
                field_name = field.get("fieldName")
                resolved_field_id = LABEL_TO_FIELD_ID.get(field_name, field_id)

            mapped_value = get_field_value(field)

            # IMPORTANT:
            # Null remains blank.
            if mapped_value is None:
                continue

            column = excel_column_for_field(resolved_field_id)

            if column is None:
                print(
                    f"WARNING: No Excel column mapping for "
                    f"'{field_id}' (fieldName='{field.get('fieldName')}')"
                )
                continue

            ws.cell(
                row=excel_row,
                column=column
            ).value = mapped_value

            total_mapped += 1

    return total_mapped


# ============================================================
# MAIN MAPPING
# ============================================================

def map_json_to_excel():
    print("=" * 70)
    print("SSI JSON -> EXCEL MAPPING")
    print("=" * 70)

    # --------------------------------------------------------
    # Validate inputs
    # --------------------------------------------------------

    if not os.path.exists(INPUT_EXCEL):
        raise FileNotFoundError(
            f"Input Excel not found:\n{INPUT_EXCEL}"
        )

    if not os.path.isdir(JSON_FOLDER):
        raise FileNotFoundError(
            f"JSON folder not found:\n{JSON_FOLDER}"
        )

    json_paths = sorted(
        glob.glob(os.path.join(JSON_FOLDER, "*.json"))
    )

    if not json_paths:
        raise FileNotFoundError(
            f"No JSON files found in:\n{JSON_FOLDER}"
        )

    print(f"JSON files found   : {len(json_paths)}")

    # --------------------------------------------------------
    # Load Excel
    # --------------------------------------------------------

    wb = load_workbook(INPUT_EXCEL)

    print(f"Excel sheets       : {wb.sheetnames}")

    processed_sheets = []

    # --------------------------------------------------------
    # Process each JSON file against its matching sheet
    # --------------------------------------------------------

    for json_path in json_paths:

        json_filename = os.path.basename(json_path)

        match = JSON_NAME_PATTERN.match(json_filename)

        if not match:
            print(
                f"WARNING: Skipping '{json_filename}' "
                f"(does not match expected naming pattern)"
            )
            continue

        doc_key = match.group(1)

        sheet_name = find_matching_sheet(doc_key, wb.sheetnames)

        if sheet_name is None:
            print(
                f"WARNING: No matching sheet for '{json_filename}' "
                f"(document key '{doc_key}')"
            )
            continue

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        records = get_records_from_json(data)

        ws = wb[sheet_name]

        total_mapped = map_records_to_sheet(ws, records)

        print(
            f"Mapped '{json_filename}' -> sheet '{sheet_name}' "
            f"(records={len(records)}, values mapped={total_mapped})"
        )

        processed_sheets.append(sheet_name)

    if not processed_sheets:
        raise ValueError(
            "No JSON files were mapped to any sheet. "
            "Check JSON_NAME_PATTERN and sheet names."
        )

    # --------------------------------------------------------
    # Save
    # --------------------------------------------------------

    wb.save(OUTPUT_EXCEL)

    print()
    print(f"Output saved       : {OUTPUT_EXCEL}")
    print(f"Sheets updated     : {processed_sheets}")

    # --------------------------------------------------------
    # Verification
    # --------------------------------------------------------

    verify_workbook(OUTPUT_EXCEL, processed_sheets)


# ============================================================
# VERIFICATION
# ============================================================

def verify_workbook(output_file, sheet_names):

    print()
    print("=" * 70)
    print("VERIFYING OUTPUT")
    print("=" * 70)

    wb = load_workbook(output_file, data_only=True)

    print(f"Sheets              : {wb.sheetnames}")

    for sheet_name in sheet_names:

        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Verification failed: '{sheet_name}' not found."
            )

        ws = wb[sheet_name]

        non_empty = []

        for row in ws.iter_rows(
            min_row=4,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column
        ):
            for cell in row:
                if cell.value is not None:
                    non_empty.append(
                        (cell.coordinate, cell.value)
                    )

        print(
            f"Non-empty cells in '{sheet_name}': "
            f"{len(non_empty)}"
        )

    print()
    print("Verification completed successfully.")


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":
    map_json_to_excel()